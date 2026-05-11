from __future__ import annotations

import csv
import json
import math
from collections import Counter
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
SOURCE_GLOB = "*.xls*"
PI = math.pi


GRID_BASES: dict[str, tuple[float, float]] = {
    "A": (170000, 2750000),
    "B": (250000, 2750000),
    "C": (330000, 2750000),
    "D": (170000, 2700000),
    "E": (250000, 2700000),
    "F": (330000, 2700000),
    "G": (170000, 2650000),
    "H": (250000, 2650000),
    "J": (90000, 2600000),
    "K": (170000, 2600000),
    "L": (250000, 2600000),
    "M": (90000, 2550000),
    "N": (170000, 2550000),
    "O": (250000, 2550000),
    "P": (90000, 2500000),
    "Q": (170000, 2500000),
    "R": (250000, 2500000),
    "T": (170000, 2450000),
    "U": (250000, 2450000),
    "V": (170000, 2400000),
    "W": (250000, 2400000),
    "X": (275000, 2614000),
    "Y": (275000, 2564000),
}


def val(text: str) -> float:
    digits = []
    for char in str(text):
        if char.isdigit() or char in ".-+":
            digits.append(char)
        else:
            break
    try:
        return float("".join(digits)) if digits else 0.0
    except ValueError:
        return 0.0


def tm2_to_lat_lng(x: float, y: float, a: float, b: float) -> tuple[float, float]:
    lng0 = 121.0 * PI / 180.0
    k0 = 0.9999
    x -= 250000.0

    e2 = 1.0 - (b**2.0) / (a**2.0)
    e1 = (1.0 - math.sqrt(1.0 - e2)) / (1.0 + math.sqrt(1.0 - e2))

    m_val = y / k0
    mu = m_val / (
        a * (1.0 - e2 / 4.0 - 3.0 * (e2**2.0) / 64.0 - 5.0 * (e2**3.0) / 256.0)
    )

    j1 = 3.0 * e1 / 2.0 - 27.0 * (e1**3.0) / 32.0
    j2 = 21.0 * (e1**2.0) / 16.0 - 55.0 * (e1**4.0) / 32.0
    j3 = 151.0 * (e1**3.0) / 96.0
    j4 = 1097.0 * (e1**4.0) / 512.0

    fp = mu + j1 * math.sin(2.0 * mu) + j2 * math.sin(4.0 * mu) + j3 * math.sin(6.0 * mu) + j4 * math.sin(8.0 * mu)

    ep2 = ((a**2.0) - (b**2.0)) / (b**2.0)
    c1 = ep2 * (math.cos(fp) ** 2.0)
    t1 = math.tan(fp) ** 2.0
    r1 = a * (1.0 - e2) / ((1.0 - e2 * (math.sin(fp) ** 2.0)) ** 1.5)
    n1 = a / math.sqrt(1.0 - e2 * (math.sin(fp) ** 2.0))
    d_val = x / (n1 * k0)

    q1 = n1 * math.tan(fp) / r1
    q2 = (d_val**2.0) / 2.0
    q3 = (5.0 + 3.0 * t1 + 10.0 * c1 - 4.0 * (c1**2.0) - 9.0 * ep2) * (d_val**4.0) / 24.0
    q4 = (61.0 + 90.0 * t1 + 298.0 * c1 + 45.0 * (t1**2.0) - 3.0 * (c1**2.0) - 252.0 * ep2) * (d_val**6.0) / 720.0

    lat = fp - q1 * (q2 - q3 + q4)
    q5 = d_val
    q6 = (1.0 + 2.0 * t1 + c1) * (d_val**3.0) / 6.0
    q7 = (5.0 - 2.0 * c1 + 28.0 * t1 - 3.0 * (c1**2.0) + 8.0 * ep2 + 24.0 * (t1**2.0)) * (d_val**5.0) / 120.0
    lng = lng0 + (q5 - q6 + q7) / math.cos(fp)

    return lat, lng


def lat_lng_to_cartesian(lat: float, lng: float, h: float, a: float, b: float) -> tuple[float, float, float]:
    e2 = 1.0 - (b**2.0) / (a**2.0)
    n = a / math.sqrt(1.0 - e2 * (math.sin(lat) ** 2.0))
    x = (n + h) * math.cos(lat) * math.cos(lng)
    y = (n + h) * math.cos(lat) * math.sin(lng)
    z = (n * (1.0 - e2) + h) * math.sin(lat)
    return x, y, z


def cartesian_to_lat_lng(x: float, y: float, z: float, a: float, b: float) -> tuple[float, float]:
    e2 = 1.0 - (b**2.0) / (a**2.0)
    ep2 = ((a**2.0) - (b**2.0)) / (b**2.0)
    p = math.sqrt(x**2.0 + y**2.0)
    theta = math.atan2(z * a, p * b)
    lat = math.atan2(z + ep2 * b * (math.sin(theta) ** 3.0), p - e2 * a * (math.cos(theta) ** 3.0))
    lng = math.atan2(y, x)
    return lat * 180.0 / PI, lng * 180.0 / PI


def tpc_to_wgs84(raw_code: object) -> tuple[float, float] | None:
    tpc_code = str(raw_code or "").strip().upper().replace(" ", "")
    if len(tpc_code) < 9:
        return None

    base = GRID_BASES.get(tpc_code[0])
    if not base:
        return None

    try:
        t2x = val(tpc_code[1:3]) * 800.0
        t2y = val(tpc_code[3:5]) * 500.0
        t3x = (ord(tpc_code[5]) - 65) * 100.0
        t3y = (ord(tpc_code[6]) - 65) * 100.0
        t99x = val(tpc_code[9:10]) if len(tpc_code) >= 11 else 0.0
        t99y = val(tpc_code[10:11]) if len(tpc_code) >= 11 else 0.0
        t5x = val(tpc_code[7:8]) * 10.0 + t99x
        t5y = val(tpc_code[8:9]) * 10.0 + t99y
    except (IndexError, TypeError):
        return None

    twd67_x = base[0] + t2x + t3x + t5x
    twd67_y = base[1] + t2y + t3y + t5y

    a_twd67 = 6378160.0
    b_twd67 = 6356774.7192
    lat67, lng67 = tm2_to_lat_lng(twd67_x, twd67_y, a_twd67, b_twd67)
    x67, y67, z67 = lat_lng_to_cartesian(lat67, lng67, 0.0, a_twd67, b_twd67)

    dx = -752.0
    dy = -358.0
    dz = -179.0
    rx = -0.0000011698
    ry = 0.0000018398
    rz = 0.0000009822
    s = 0.00002329

    x84 = x67 + dx + s * x67 - rz * y67 + ry * z67
    y84 = y67 + dy + rz * x67 + s * y67 - rx * z67
    z84 = z67 + dz - ry * x67 + rx * y67 + s * z67

    a_wgs84 = 6378137.0
    b_wgs84 = 6356752.314245
    return cartesian_to_lat_lng(x84, y84, z84, a_wgs84, b_wgs84)


def find_excel_file() -> Path:
    files = [path for path in ROOT.glob(SOURCE_GLOB) if not path.name.startswith("~$")]
    if not files:
        raise FileNotFoundError("找不到 Excel 檔案")
    return max(files, key=lambda path: path.stat().st_mtime)


def normalize_header(value: object) -> str:
    return str(value or "").strip().replace(" ", "")


def find_headers(headers: Iterable[object]) -> dict[str, int]:
    wanted = {"土木設備": "device", "圖號座標": "code", "區域": "area"}
    found: dict[str, int] = {}
    for index, header in enumerate(headers):
        key = wanted.get(normalize_header(header))
        if key:
            found[key] = index
    missing = [name for name in wanted.values() if name not in found]
    if missing:
        raise ValueError(f"Excel 欄位缺少: {', '.join(missing)}")
    return found


def main() -> None:
    DATA_DIR.mkdir(exist_ok=True)
    source = find_excel_file()
    points_path = DATA_DIR / "points.json"
    csv_path = DATA_DIR / "points.csv"
    meta_path = DATA_DIR / "meta.json"

    workbook = load_workbook(source, read_only=True, data_only=True, keep_vba=False)
    sheet = workbook.worksheets[0]
    rows = sheet.iter_rows(values_only=True)
    headers = next(rows)
    indexes = find_headers(headers)

    converted = 0
    skipped = 0
    min_lat = 90.0
    min_lng = 180.0
    max_lat = -90.0
    max_lng = -180.0
    area_counter: Counter[str] = Counter()

    with points_path.open("w", encoding="utf-8") as json_file, csv_path.open("w", encoding="utf-8-sig", newline="") as csv_file:
        writer = csv.writer(csv_file)
        writer.writerow(["id", "name", "code", "area", "lat", "lng"])
        json_file.write("[\n")
        first = True

        for row_number, row in enumerate(rows, start=2):
            device = str(row[indexes["device"]] or "").strip()
            code = str(row[indexes["code"]] or "").strip()
            area = str(row[indexes["area"]] or "").strip()
            result = tpc_to_wgs84(code)
            if not result:
                skipped += 1
                continue

            lat, lng = result
            if not (20.0 <= lat <= 27.0 and 118.0 <= lng <= 123.5):
                skipped += 1
                continue

            converted += 1
            min_lat = min(min_lat, lat)
            min_lng = min(min_lng, lng)
            max_lat = max(max_lat, lat)
            max_lng = max(max_lng, lng)
            area_counter[area] += 1

            record = {
                "id": converted,
                "name": device,
                "code": code,
                "area": area,
                "lat": round(lat, 7),
                "lng": round(lng, 7),
            }
            writer.writerow([record["id"], record["name"], record["code"], record["area"], record["lat"], record["lng"]])
            if not first:
                json_file.write(",\n")
            json.dump(record, json_file, ensure_ascii=False, separators=(",", ":"))
            first = False

            if converted % 50000 == 0:
                print(f"converted {converted:,} rows...")

        json_file.write("\n]\n")

    workbook.close()

    meta = {
        "source": source.name,
        "totalRows": sheet.max_row - 1,
        "converted": converted,
        "skipped": skipped,
        "bounds": [[min_lat, min_lng], [max_lat, max_lng]],
        "areas": [{"name": name, "count": count} for name, count in area_counter.most_common()],
    }
    meta_path.write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(meta, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
